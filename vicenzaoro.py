"""
Vicenzaoro Catalogue Scraper — FIXED VERSION
=============================================
API behaviour (verified):
  - POST /en/api/v1/filterDataObjects
  - Always returns exactly 10 items per page (perPage param is ignored by server)
  - Response structure:  {"itemsCount": 10, "items": [...], "paginationVariables": {...}}
  - Each item["content"] is an HTML string of the card
  - When page is out of range the server returns an empty list:  []
  - Total exhibitors ≈ 1278  (128 pages × 10 + 8 on last page)

Requirements:
    pip install requests beautifulsoup4 openpyxl tqdm
"""

import re
import time
import random
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm

# ─────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────
BASE_URL     = "https://www.vicenzaoro.com"
CATALOGUE    = f"{BASE_URL}/en/catalogue"
API_URL      = f"{BASE_URL}/en/api/v1/filterDataObjects"
OUTPUT_FILE  = "vicenzaoro_exhibitors.xlsx"
MAX_WORKERS  = 25        # parallel profile-page threads (≈ 25 browser tabs)
DELAY_MIN    = 0.2       # polite delay per thread (seconds)
DELAY_MAX    = 0.6

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/html, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Content-Type":    "application/json",
    "Referer":         CATALOGUE,
}


# ─────────────────────────────────────────────────────────────
# STEP 1  —  Collect ALL exhibitor links via the paginated API
# ─────────────────────────────────────────────────────────────
def get_all_exhibitor_links() -> list[dict]:
    """
    Pages through the internal API until it returns an empty response.
    Returns list of  {name, position, url, highlighted}
    """
    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Read folder / edition IDs from the live catalogue page
    print("[*] Loading catalogue page to read API parameters …")
    resp = session.get(CATALOGUE, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    listing = soup.find(class_="listing-data-object")
    if not listing:
        raise RuntimeError("Could not find .listing-data-object on the catalogue page.")

    folder  = listing.get("data-folder", "4656479")
    edition = listing.get("data-exhibition-edition", "340271")
    print(f"    folder={folder}  edition={edition}")

    all_items: list[dict] = []
    page_num = 1

    while True:
        payload = {
            "type":              "digital-profiles",
            "folder":            folder,
            "exhibitionEdition": edition,
            "page":              page_num,
            "perPage":           10,       # server always returns 10 regardless
            "locale":            "en",
        }

        print(f"[*] API page {page_num:>4} …", end=" ", flush=True)

        try:
            api_resp = session.post(API_URL, json=payload, timeout=30)
            api_resp.raise_for_status()
            data = api_resp.json()
        except Exception as exc:
            print(f"  ✗ Request failed: {exc} — stopping.")
            break

        # ── STOP CONDITION: server returns [] when page is out of range
        if isinstance(data, list):
            if len(data) == 0:
                print("  empty list — end of pages.")
            else:
                print(f"  unexpected list of {len(data)} — stopping.")
            break

        # ── Normal dict response
        if not isinstance(data, dict):
            print(f"  unexpected type {type(data)} — stopping.")
            break

        items_count = data.get("itemsCount", 0)
        raw_items   = data.get("items", [])

        print(f"  {items_count} items")

        if not raw_items or items_count == 0:
            print("  No items — end of pagination.")
            break

        # ── Parse each item's HTML content fragment
        parsed = _parse_items(raw_items)
        all_items.extend(parsed)

        page_num += 1
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    # Deduplicate by URL (keeps first occurrence)
    seen, unique = set(), []
    for item in all_items:
        if item["url"] and item["url"] not in seen:
            seen.add(item["url"])
            unique.append(item)

    print(f"\n[✓] Collected {len(unique)} unique exhibitor profiles.")
    return unique


def _parse_items(raw_items: list) -> list[dict]:
    """
    Each item is a dict whose 'content' key holds an HTML card string.
    Extract name, hall/position, profile URL, and highlighted flag.
    """
    results = []
    for item in raw_items:
        content = item.get("content", "")
        if not content:
            continue

        frag = BeautifulSoup(content, "html.parser")

        name_el = frag.find(class_="card-digitalprofile-name")
        pos_el  = frag.find(class_="card-digitalprofile-position")
        link_el = frag.find("a", href=re.compile(r"profile-detail"))

        if not link_el:
            continue

        name = name_el.get_text(strip=True) if name_el else ""

        # Clean "POSITION  H 4 /121" → "H 4 /121"
        raw_pos = pos_el.get_text(" ", strip=True) if pos_el else ""
        position = re.sub(r"(?i)^position\s*", "", raw_pos).strip()

        href = link_el.get("href", "")
        url  = href if href.startswith("http") else BASE_URL + href

        highlighted = bool(frag.find(class_=re.compile(r"\bhighlighted\b")))

        results.append({
            "name":        name,
            "position":    position,
            "url":         url,
            "highlighted": highlighted,
        })
    return results


# ─────────────────────────────────────────────────────────────
# STEP 2  —  Scrape each exhibitor profile page
# ─────────────────────────────────────────────────────────────
def scrape_profile(item: dict) -> dict:
    """Fetch one profile page and extract all available fields."""
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    record = {
        "Name":          item.get("name", ""),
        "Hall / Position": item.get("position", ""),
        "Highlighted":   "Yes" if item.get("highlighted") else "No",
        "Address":       "",
        "Phone":         "",
        "Website":       "",
        "Email":         "",
        "Social":        "",
        "Categories":    "",
        "Sub-categories":"",
        "Description":   "",
        "Products":      "",
        "Profile URL":   item.get("url", ""),
    }

    try:
        resp = requests.get(item["url"], headers=HEADERS, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # ── Company name (h1)
        h1 = soup.find("h1")
        if h1:
            record["Name"] = h1.get_text(strip=True)

        # ── Hall / Position (sibling div under the heading block)
        for el in soup.find_all(string=True):
            stripped = el.strip()
            if re.match(r"^H\s*[\d\.]+[A-Z]?\s*/", stripped, re.IGNORECASE):
                record["Hall / Position"] = stripped
                break

        # ── Description paragraph
        desc_candidates = soup.find_all("p")
        for p in desc_candidates:
            txt = p.get_text(strip=True)
            if len(txt) > 40 and "register" not in txt.lower():
                record["Description"] = txt
                break

        # ── REFERENCES, WEBSITE, EMAIL, SOCIAL via section labels
        _extract_labelled_sections(soup, record)

        # ── Categories / sub-categories (tag pills)
        _extract_categories(soup, record)

        # ── Product names
        products = []
        for prod in soup.find_all(class_=re.compile(r"product", re.I)):
            txt = prod.get_text(strip=True)
            if txt and len(txt) < 100:
                products.append(txt)
        # Deduplicate while preserving order
        seen_p: set[str] = set()
        unique_products = []
        for p in products:
            if p not in seen_p:
                seen_p.add(p)
                unique_products.append(p)
        record["Products"] = " | ".join(unique_products[:15])

    except Exception as exc:
        record["Description"] = f"ERROR: {exc}"

    return record


def _extract_labelled_sections(soup: BeautifulSoup, record: dict):
    """
    Walk text nodes looking for section headers like REFERENCES, WEBSITE,
    EMAIL, SOCIAL and grab the surrounding block's content.
    """
    LABELS = {"REFERENCES", "WEBSITE", "EMAIL", "SOCIAL"}

    for text_node in soup.find_all(string=True):
        label = text_node.strip().upper()
        if label not in LABELS:
            continue

        parent = text_node.parent
        if not parent:
            continue
        block = parent.parent  # one level up contains the full section block
        if not block:
            block = parent

        if label == "REFERENCES":
            lines = [
                ln.strip()
                for ln in block.get_text("\n").splitlines()
                if ln.strip() and ln.strip().upper() != "REFERENCES"
            ]
            phone_lines = [ln for ln in lines if re.search(r"tel\.?:|phone", ln, re.I)]
            addr_lines  = [ln for ln in lines if ln not in phone_lines]
            if addr_lines:
                record["Address"] = ", ".join(addr_lines)
            if phone_lines:
                record["Phone"] = re.sub(r"(?i)tel\.?:\s*", "", phone_lines[0]).strip()

        elif label == "WEBSITE":
            link = block.find("a", href=re.compile(r"^https?://"))
            if link:
                href = link["href"]
                if "vicenzaoro" not in href and "iegexpo" not in href:
                    record["Website"] = href
            else:
                # Plain text URL
                for ln in block.get_text("\n").splitlines():
                    ln = ln.strip()
                    if ln.startswith("http") and "vicenzaoro" not in ln:
                        record["Website"] = ln
                        break

        elif label == "EMAIL":
            link = block.find("a", href=re.compile(r"^mailto:"))
            if link:
                record["Email"] = link["href"].replace("mailto:", "").strip()
            else:
                for ln in block.get_text("\n").splitlines():
                    ln = ln.strip()
                    if "@" in ln and "EMAIL" not in ln.upper():
                        record["Email"] = ln
                        break

        elif label == "SOCIAL":
            social_links = [
                a["href"]
                for a in block.find_all("a", href=re.compile(
                    r"facebook|instagram|linkedin|twitter|youtube|tiktok", re.I
                ))
                if a.get("href")
            ]
            record["Social"] = " | ".join(social_links)


def _extract_categories(soup: BeautifulSoup, record: dict):
    """Extract main categories and sub-category tag pills."""
    main_cats: list[str] = []
    sub_cats:  list[str] = []

    # Look for the Categories section
    cat_section = soup.find(string=re.compile(r"^Categories$", re.I))
    if cat_section:
        block = cat_section.parent
        # Walk up to find the full categories block
        for _ in range(4):
            if block and len(block.find_all(class_=re.compile(r"tag|categor|pill", re.I))) > 0:
                break
            if block:
                block = block.parent

        if block:
            # Main category headings (uppercase, not tag pills)
            for el in block.find_all(True):
                txt = el.get_text(strip=True)
                # Main categories tend to be short uppercase strings without pills
                if (el.name in ("strong", "b", "h4", "h5", "p", "div")
                        and txt.isupper()
                        and 3 < len(txt) < 60
                        and txt != "CATEGORIES"):
                    if txt not in main_cats:
                        main_cats.append(txt)

            # Sub-category tags (pill/tag spans)
            for tag in block.find_all(class_=re.compile(r"tag|pill|badge", re.I)):
                txt = tag.get_text(strip=True)
                if txt and len(txt) < 80 and txt not in sub_cats:
                    sub_cats.append(txt)

    record["Categories"]     = " | ".join(main_cats)
    record["Sub-categories"] = " | ".join(sub_cats)


# ─────────────────────────────────────────────────────────────
# STEP 3  —  Write to Excel
# ─────────────────────────────────────────────────────────────
COLUMNS = [
    "Name", "Hall / Position", "Highlighted", "Address", "Phone",
    "Website", "Email", "Social", "Categories", "Sub-categories",
    "Description", "Products", "Profile URL",
]

COL_WIDTHS = {
    "Name": 32, "Hall / Position": 16, "Highlighted": 12,
    "Address": 38, "Phone": 22, "Website": 38, "Email": 32,
    "Social": 45, "Categories": 35, "Sub-categories": 45,
    "Description": 65, "Products": 45, "Profile URL": 48,
}

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
ALT_FILL    = PatternFill("solid", fgColor="EDF1F8")


def write_excel(records: list[dict], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vicenzaoro Exhibitors"

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Data rows
    for row_idx, rec in enumerate(records, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val  = rec.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Clickable hyperlinks
            if col_name in ("Website", "Profile URL") and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="1155CC", underline="single")

            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Column widths
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = COL_WIDTHS.get(col_name, 20)

    ws.freeze_panes = "A2"       # keep header visible while scrolling
    ws.auto_filter.ref = ws.dimensions   # enable column filters

    wb.save(filename)
    print(f"\n[✓] Saved → {filename}  ({len(records)} rows, {len(COLUMNS)} columns)")


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    print("=" * 62)
    print("  Vicenzaoro Catalogue Scraper  —  FIXED")
    print("=" * 62)

    # 1 ── Collect all exhibitor links (paginate through API)
    exhibitors = get_all_exhibitor_links()
    if not exhibitors:
        print("[!] No exhibitors found — check your connection or folder/edition IDs.")
        return

    # 2 ── Scrape each profile with up to 25 parallel threads
    print(f"\n[*] Scraping {len(exhibitors)} profiles ({MAX_WORKERS} threads) …\n")
    records: list[dict] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scrape_profile, ex): ex for ex in exhibitors}
        for future in tqdm(as_completed(futures), total=len(futures), unit="profile"):
            try:
                records.append(future.result())
            except Exception as exc:
                ex = futures[future]
                print(f"\n  ✗ {ex.get('name','?')} — {exc}")

    # Sort alphabetically
    records.sort(key=lambda r: r.get("Name", "").upper())

    # 3 ── Export
    write_excel(records, OUTPUT_FILE)

    # Quick preview
    print("\n── Preview (first 4 rows) ──")
    print(f"{'Name':<35} {'Hall':<15} {'Email':<30} {'Website'}")
    print("-" * 100)
    for r in records[:4]:
        print(f"{r['Name']:<35} {r['Hall / Position']:<15} "
              f"{r['Email']:<30} {r['Website']}")


if __name__ == "__main__":
    main()