"""
GaLaBau Exhibitor Scraper – v2 (Algolia-based)
===============================================
Strategy:
  1. Pull ALL 960 exhibitors directly from the Algolia API in one shot (no Selenium needed).
  2. Scrape each detail page only to pick up phone, website, and social-media links
     (not available in the Algolia index), using data-testid attributes for reliability.
  3. Export to Excel with summary sheet.

Requirements:
    pip install requests openpyxl beautifulsoup4 tqdm
"""

import re
import time
import logging
import datetime
import base64
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
BASE_URL        = "https://www.galabau-messe.com"
OUTPUT_FILE     = "galabau_exhibitors.xlsx"
MAX_WORKERS     = 15       # concurrent detail-page threads
REQUEST_DELAY   = 0.2      # seconds between HTTP requests per thread
HITS_PER_PAGE   = 100      # Algolia allows up to 1000; 100 is safe and fast

# ── Algolia credentials (public read-only search key found in page JS) ──
ALGOLIA_APP_ID   = "4EB6G0V1NT"
ALGOLIA_API_KEY  = "f0416e3d1b38ae3aa789c8750e12bfe5"
ALGOLIA_INDEX    = "prod_website_companies_en"
ALGOLIA_FILTER   = "site:gala"
ALGOLIA_ENDPOINT = f"https://{ALGOLIA_APP_ID}-dsn.algolia.net/1/indexes/*/queries"

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# DATA MODEL
# ──────────────────────────────────────────────
@dataclass
class Exhibitor:
    name:             str = ""
    url:              str = ""
    hall:             str = ""
    booth_number:     str = ""
    company_type:     str = ""
    country:          str = ""
    address:          str = ""
    postal_code:      str = ""
    city:             str = ""
    phone:            str = ""
    website:          str = ""
    email:            str = ""
    about_us:         str = ""
    products_offered: str = ""
    keywords:         str = ""
    tagline:          str = ""
    social_youtube:   str = ""
    social_instagram: str = ""
    social_facebook:  str = ""
    social_linkedin:  str = ""
    social_twitter:   str = ""
    logo_url:         str = ""
    employees:        str = ""   # "FirstName LastName <email>" joined by " | "


# ──────────────────────────────────────────────
# STEP 1 – Pull all records from Algolia API
# ──────────────────────────────────────────────
def _algolia_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "X-Algolia-Application-Id": ALGOLIA_APP_ID,
        "X-Algolia-API-Key":        ALGOLIA_API_KEY,
        "Content-Type":             "application/json",
    })
    return s


def fetch_all_from_algolia() -> list[Exhibitor]:
    """
    Paginate through the Algolia index and return one Exhibitor per hit.
    All fields available in the index are populated here.
    """
    session = _algolia_session()

    # First call: get total number of hits
    payload = {
        "requests": [{
            "indexName":   ALGOLIA_INDEX,
            "filters":     ALGOLIA_FILTER,
            "hitsPerPage": 1,
            "page":        0,
            "query":       "",
        }]
    }
    resp = session.post(ALGOLIA_ENDPOINT, json=payload, timeout=20)
    resp.raise_for_status()
    total_hits = resp.json()["results"][0]["nbHits"]
    num_pages   = -(-total_hits // HITS_PER_PAGE)   # ceiling division
    log.info(f"Algolia reports {total_hits} exhibitors across {num_pages} pages "
             f"({HITS_PER_PAGE} per page).")

    all_hits: list[dict] = []
    for page in range(num_pages):
        payload["requests"][0].update({"hitsPerPage": HITS_PER_PAGE, "page": page})
        r = session.post(ALGOLIA_ENDPOINT, json=payload, timeout=20)
        r.raise_for_status()
        hits = r.json()["results"][0]["hits"]
        all_hits.extend(hits)
        log.info(f"  Page {page + 1}/{num_pages} – {len(all_hits)} records fetched so far.")
        time.sleep(0.1)   # be polite

    log.info(f"Total records fetched from Algolia: {len(all_hits)}")
    return [_hit_to_exhibitor(h) for h in all_hits]


def _hit_to_exhibitor(hit: dict) -> Exhibitor:
    """Map one Algolia hit dict → Exhibitor dataclass."""
    ex = Exhibitor()

    ex.name       = hit.get("companyName", "") or ""
    ex.tagline    = hit.get("slogan", "")      or ""
    ex.country    = hit.get("country", "")     or ""
    ex.address    = hit.get("streetno", "")    or ""
    ex.postal_code= hit.get("postcode", "")    or ""
    ex.city       = hit.get("city", "")        or ""
    ex.company_type = hit.get("companyType", "") or ""
    ex.logo_url   = hit.get("logo", "")        or ""
    ex.email      = hit.get("email", "")       or ""

    # Relative URL → absolute
    raw_url = hit.get("url", "") or ""
    ex.url  = (BASE_URL + raw_url) if raw_url.startswith("/") else raw_url

    # Hall & booth (first booth entry)
    booths = hit.get("booth") or []
    if booths:
        ex.hall         = booths[0].get("boothHall", "")   or ""
        ex.booth_number = booths[0].get("boothNumber", "") or ""

    # About us (HTML → plain text)
    desc = hit.get("companyDescription", "") or ""
    if desc:
        soup = BeautifulSoup(desc, "html.parser")
        ex.about_us = soup.get_text(" ", strip=True)

    # Keywords (array → pipe-separated)
    kw = hit.get("keyword") or []
    if isinstance(kw, list):
        ex.keywords = " | ".join(str(k).strip() for k in kw if k)
    else:
        ex.keywords = str(kw)

    # Products offered (array → pipe-separated)
    prods = hit.get("products") or []
    if isinstance(prods, list):
        ex.products_offered = " | ".join(str(p).strip() for p in prods if p)
    else:
        ex.products_offered = str(prods)

    # Employees (list of dicts)
    employees = hit.get("employee") or []
    parts = []
    for emp in employees:
        fn = emp.get("firstName", "") or ""
        ln = emp.get("lastName", "")  or ""
        em = emp.get("email", "")     or ""
        entry = f"{fn} {ln}".strip()
        if em:
            entry += f" <{em}>"
        if entry:
            parts.append(entry)
    ex.employees = " | ".join(parts)

    return ex


# ──────────────────────────────────────────────
# STEP 2 – Enrich with phone / website / social
#           from each exhibitor's detail page
# ──────────────────────────────────────────────
DETAIL_SESSION = requests.Session()
DETAIL_SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
})


def _decode_b64_email(href: str) -> str:
    """
    The site encodes emails as base64 in mailto: hrefs.
    e.g.  mailto:dy5tdWVsbGVyQGZlcmF4LmRl  →  w.mueller@ferax.de
    """
    raw = href.replace("mailto:", "")
    # Check if it looks like base64 (no @ sign = encoded)
    if "@" in raw:
        return raw          # already plain text
    try:
        padded = raw + "=" * (-len(raw) % 4)
        return base64.b64decode(padded).decode("utf-8")
    except Exception:
        return raw


def enrich_from_detail_page(ex: Exhibitor) -> Exhibitor:
    """
    Fetch the exhibitor detail page and fill in:
      phone, website, social media links
    Uses data-testid attributes for robust, class-independent extraction.
    """
    if not ex.url:
        return ex

    time.sleep(REQUEST_DELAY)
    try:
        resp = DETAIL_SESSION.get(ex.url, timeout=25)
        resp.raise_for_status()
    except requests.RequestException as e:
        log.warning(f"Detail page failed for {ex.url}: {e}")
        return ex

    soup = BeautifulSoup(resp.text, "html.parser")

    def by_testid(tid: str):
        return soup.find(attrs={"data-testid": tid})

    # ── Website ──────────────────────────────
    ws_el = by_testid("company-details-contacts-website")
    if ws_el and ws_el.name == "a":
        ex.website = ws_el.get("href", "").strip()

    # ── Phone ────────────────────────────────
    ph_el = by_testid("company-details-contacts-phone")
    if ph_el and ph_el.name == "a":
        href = ph_el.get("href", "")
        ex.phone = href.replace("tel:", "").strip() if href.startswith("tel:") \
                   else ph_el.get_text(strip=True)

    # ── Email (if not already from Algolia) ──
    if not ex.email:
        em_el = by_testid("company-details-contacts-email")
        if em_el and em_el.name == "a":
            ex.email = _decode_b64_email(em_el.get("href", ""))

    # ── Social media ─────────────────────────
    social_testids = {
        "company-details-contacts-youtube":   "social_youtube",
        "company-details-contacts-instagram": "social_instagram",
        "company-details-contacts-facebook":  "social_facebook",
        "company-details-contacts-linkedin":  "social_linkedin",
        "company-details-contacts-twitter":   "social_twitter",
        "company-details-contacts-xing":      "social_twitter",  # sometimes Xing
    }
    for tid, attr in social_testids.items():
        el = by_testid(tid)
        if el and el.name == "a":
            setattr(ex, attr, el.get("href", "").strip())

    # Fallback: scan all <a> tags for social platforms not caught by testid
    if not any([ex.social_youtube, ex.social_instagram, ex.social_facebook,
                ex.social_linkedin, ex.social_twitter]):
        platform_map = {
            "youtube.com":  "social_youtube",
            "instagram.com":"social_instagram",
            "facebook.com": "social_facebook",
            "linkedin.com": "social_linkedin",
            "twitter.com":  "social_twitter",
            "x.com":        "social_twitter",
        }
        for a in soup.find_all("a", href=True):
            href = a["href"].lower()
            for platform, field in platform_map.items():
                if platform in href and not getattr(ex, field):
                    setattr(ex, field, a["href"])

    return ex


# ──────────────────────────────────────────────
# STEP 3 – Enrich all concurrently
# ──────────────────────────────────────────────
def enrich_all(exhibitors: list[Exhibitor]) -> list[Exhibitor]:
    log.info(f"Enriching {len(exhibitors)} detail pages with {MAX_WORKERS} threads …")
    results: list[Exhibitor] = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(enrich_from_detail_page, ex): ex for ex in exhibitors}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Enriching"):
            try:
                results.append(future.result())
            except Exception as e:
                log.warning(f"Unexpected error: {e}")
    results.sort(key=lambda e: e.name.lower())
    return results


# ──────────────────────────────────────────────
# STEP 4 – Export to Excel
# ──────────────────────────────────────────────
COLUMNS = [
    ("Company Name",     "name"),
    ("Hall",             "hall"),
    ("Booth Number",     "booth_number"),
    ("Company Type",     "company_type"),
    ("Country",          "country"),
    ("Address",          "address"),
    ("Postal Code",      "postal_code"),
    ("City",             "city"),
    ("Phone",            "phone"),
    ("Website",          "website"),
    ("Email",            "email"),
    ("Tagline",          "tagline"),
    ("Products Offered", "products_offered"),
    ("About Us",         "about_us"),
    ("Keywords",         "keywords"),
    ("Employees",        "employees"),
    ("YouTube",          "social_youtube"),
    ("Instagram",        "social_instagram"),
    ("Facebook",         "social_facebook"),
    ("LinkedIn",         "social_linkedin"),
    ("Twitter/X",        "social_twitter"),
    ("Logo URL",         "logo_url"),
    ("Profile URL",      "url"),
]

HEADER_FILL  = PatternFill("solid", fgColor="1D5E4C")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="E8F5E9")

COL_WIDTHS = {
    "Company Name": 35, "Hall": 10, "Booth Number": 15, "Company Type": 20,
    "Country": 18, "Address": 30, "Postal Code": 12, "City": 20,
    "Phone": 20, "Website": 35, "Email": 35, "Tagline": 40,
    "Products Offered": 40, "About Us": 60, "Keywords": 50, "Employees": 45,
    "YouTube": 35, "Instagram": 35, "Facebook": 35, "LinkedIn": 35,
    "Twitter/X": 35, "Logo URL": 40, "Profile URL": 55,
}


def export_to_excel(data: list[Exhibitor], filepath: str) -> None:
    log.info(f"Exporting {len(data)} records to {filepath} …")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhibitors"

    # Header row
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, exhibitor in enumerate(data, start=2):
        for col_idx, (_, attr) in enumerate(COLUMNS, start=1):
            value = getattr(exhibitor, attr, "") or ""
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Column widths
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = COL_WIDTHS.get(header, 20)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "GaLaBau Exhibitor Scrape Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"]  = "Total Exhibitors:"   ; ws2["B3"]  = len(data)
    ws2["A4"]  = "With Website:"       ; ws2["B4"]  = sum(1 for e in data if e.website)
    ws2["A5"]  = "With Email:"         ; ws2["B5"]  = sum(1 for e in data if e.email)
    ws2["A6"]  = "With Phone:"         ; ws2["B6"]  = sum(1 for e in data if e.phone)
    ws2["A7"]  = "With About Us:"      ; ws2["B7"]  = sum(1 for e in data if e.about_us)
    ws2["A8"]  = "With Keywords:"      ; ws2["B8"]  = sum(1 for e in data if e.keywords)
    ws2["A9"]  = "With Social Media:"  ; ws2["B9"]  = sum(
        1 for e in data if any([e.social_youtube, e.social_instagram,
                                 e.social_facebook, e.social_linkedin, e.social_twitter]))
    ws2["A11"] = "Source URL:"         ; ws2["B11"] = "https://www.galabau-messe.com/en/exhibitors-products/find-exhibitors"
    ws2["A12"] = "Scraped on:"         ; ws2["B12"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 55

    wb.save(filepath)
    log.info(f"✅  Saved: {filepath}")


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def main():
    # 1. Bulk-fetch all exhibitors from Algolia (no Selenium, no browser needed)
    exhibitors = fetch_all_from_algolia()
    if not exhibitors:
        log.error("No exhibitors fetched – aborting.")
        return

    # 2. Enrich each with phone/website/social from detail pages (concurrent)
    exhibitors = enrich_all(exhibitors)

    # 3. Export
    export_to_excel(exhibitors, OUTPUT_FILE)
    log.info(f"\nDone!  {len(exhibitors)} exhibitors saved to '{OUTPUT_FILE}'")


if __name__ == "__main__":
    main()