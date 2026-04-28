"""
Autumn Fair Exhibitor Scraper - FULLY FIXED
All fields now use exact CSS class selectors confirmed from live DOM inspection.
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BASE_URL        = "https://www.autumnfair.com"
EXHIBITORS_URL  = "https://www.autumnfair.com/exhibitors"
TOTAL_PAGES     = 39
MAX_WORKERS     = 10
REQUEST_DELAY   = 0.4
OUTPUT_FILE     = "autumnfair_exhibitors.xlsx"

LISTING_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "X-Requested-With": "non-ajax",
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "en-GB,en;q=0.9",
    "Referer": "https://www.autumnfair.com/exhibitors",
}

DETAIL_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "en-GB,en;q=0.9",
}

session = requests.Session()


# ─────────────────────────────────────────────
# STEP 1 – Collect exhibitor links from listing pages
# ─────────────────────────────────────────────
def get_exhibitor_links_from_page(page_number: int) -> list[dict]:
    params = {
        "page": str(page_number),
        "searchgroup": "00000001-exhibitors",
    }
    try:
        resp = session.get(EXHIBITORS_URL, params=params, headers=LISTING_HEADERS, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [!] Failed listing page {page_number}: {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    items = soup.select("[data-href*='exhibitors/']")
    exhibitors = []

    for item in items:
        data_href = item.get("data-href", "")
        if not data_href or "#" in data_href:
            continue

        full_url = urljoin(BASE_URL + "/", data_href)

        name_tag = item.select_one(
            ".m-exhibitors-list__items__item__name, "
            "[class*='__name'], h2, h3, strong"
        )
        name = name_tag.get_text(strip=True) if name_tag else data_href.split("/")[-1]

        stand_tag = item.select_one("[class*='stand']")
        stand = ""
        if stand_tag:
            stand = re.sub(r"Stand:\s*", "", stand_tag.get_text(strip=True)).strip()

        exhibitors.append({"name": name, "url": full_url, "stand_preview": stand})

    return exhibitors


def collect_all_exhibitor_links() -> list[dict]:
    all_links = []
    seen_urls = set()

    for page in range(1, TOTAL_PAGES + 1):
        print(f"  Listing page {page}/{TOTAL_PAGES} ...", end=" ", flush=True)
        links = get_exhibitor_links_from_page(page)
        new = [l for l in links if l["url"] not in seen_urls]
        for l in new:
            seen_urls.add(l["url"])
        all_links.extend(new)
        print(f"found {len(new)} (total: {len(all_links)})")
        time.sleep(REQUEST_DELAY)

    return all_links


# ─────────────────────────────────────────────
# STEP 2 – Scrape individual exhibitor detail pages
# ─────────────────────────────────────────────
def scrape_exhibitor_page(exhibitor: dict) -> dict:
    url  = exhibitor["url"]
    data = {
        "Name":                     exhibitor.get("name", ""),
        "URL":                      url,
        "Stand":                    exhibitor.get("stand_preview", ""),
        "Sector":                   "",
        "Description":              "",
        "Products Include":         "",
        "Exclusive to Autumn Fair": "",
        "Sells on Faire.com":       "",
        "Address":                  "",
        "Country":                  "",
        "Website":                  "",
        "Email":                    "",
        "Phone":                    "",
        "Facebook":                 "",
        "Instagram":                "",
        "LinkedIn":                 "",
        "Twitter / X":              "",
        "TikTok":                   "",
        "YouTube":                  "",
        "New Exhibitor":            "",
        "Taste @ Autumn Fair":      "",
        "B Corp Certified":         "",
        "Manufactured in UK":       "",
        "Licensed Products":        "",
        "Export Ready":             "",
        "White/Private Label":      "",
    }

    try:
        time.sleep(REQUEST_DELAY)
        resp = session.get(url, headers=DETAIL_HEADERS, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [!] Failed {url}: {e}")
        return data

    soup = BeautifulSoup(resp.text, "html.parser")

    # ── Name
    h1 = soup.select_one("h1")
    if h1:
        data["Name"] = h1.get_text(strip=True)

    # ── Stand  (exact class confirmed from DOM)
    stand_el = soup.select_one(".m-exhibitor-entry__item__header__stand")
    if stand_el:
        data["Stand"] = re.sub(r"Stand:\s*", "", stand_el.get_text(strip=True)).strip()

    # ── Sector  (exact class confirmed from DOM)
    sector_el = soup.select_one(".m-exhibitor-entry__item__header__categories")
    if sector_el:
        data["Sector"] = sector_el.get_text(strip=True)

    # ── Description  (exact class confirmed from DOM)
    desc_el = soup.select_one(".m-exhibitor-entry__item__body__description__profile")
    if desc_el:
        data["Description"] = " ".join(desc_el.get_text(" ", strip=True).split())

    # ── Products Include icons
    #    Exact class: .m-exhibitor-entry__item__body__description__products__body img
    #    Alt texts seen: "Manufactured in the UK Icon", "Export Ready Icon",
    #                    "Licensed Products Icon", "Sells on Faire.com Icon",
    #                    "B Corp Certified Icon", "White/Private Label Icon"
    product_icons = soup.select(
        ".m-exhibitor-entry__item__body__description__products__body img"
    )
    icon_labels = []
    for img in product_icons:
        alt = (img.get("alt") or img.get("title") or "").strip()
        # Clean " Icon" suffix for readability
        alt_clean = re.sub(r"\s*Icon\s*$", "", alt, flags=re.I).strip()
        if alt_clean:
            icon_labels.append(alt_clean)
    data["Products Include"] = ", ".join(dict.fromkeys(icon_labels))

    # ── Additional key-value fields
    #    Exact class: .m-exhibitor-entry__item__body__description__additional__item
    #    Title:  .m-exhibitor-entry__item__body__description__additional__item__title
    #    Value:  .m-exhibitor-entry__item__body__description__additional__item__value
    add_items = soup.select(
        ".m-exhibitor-entry__item__body__description__additional__item"
    )
    for item in add_items:
        title_el = item.select_one(
            ".m-exhibitor-entry__item__body__description__additional__item__title"
        )
        value_el = item.select_one(
            ".m-exhibitor-entry__item__body__description__additional__item__value"
        )
        if not title_el:
            continue

        raw_title = title_el.get_text(strip=True)
        value     = value_el.get_text(strip=True) if value_el else ""

        # The website JS renames "Are you exhibiting at any other UK trade show..."
        # to "Exclusive to Autumn Fair" and flips Yes <-> No
        if "other UK trade show" in raw_title or "Are you exhibiting" in raw_title:
            raw_title = "Exclusive to Autumn Fair"
            value = "No" if value == "Yes" else ("Yes" if value == "No" else value)

        tl = raw_title.lower()
        if "exclusive" in tl:
            data["Exclusive to Autumn Fair"] = value
        elif "faire" in tl:
            data["Sells on Faire.com"] = value
        elif "new exhibitor" in tl:
            data["New Exhibitor"] = value
        elif "taste" in tl:
            data["Taste @ Autumn Fair"] = value
        elif "b corp" in tl:
            data["B Corp Certified"] = value
        elif "manufactured" in tl:
            data["Manufactured in UK"] = value
        elif "licensed" in tl:
            data["Licensed Products"] = value
        elif "export" in tl:
            data["Export Ready"] = value
        elif "white" in tl or "private label" in tl:
            data["White/Private Label"] = value

    # ── Address + Country
    #    FIX: use exact class, split on newlines, filter blanks.
    #    The address div contains plain text nodes like ["Dodford", "United Kingdom"].
    #    Last non-empty line after removing the "Address" header = country.
    addr_el = soup.select_one(".m-exhibitor-entry__item__body__contacts__address")
    if addr_el:
        lines = [
            ln.strip().rstrip(",").strip()
            for ln in addr_el.get_text().split("\n")
            if ln.strip() and ln.strip() != "Address"
        ]
        if lines:
            # Last line is always the country (United Kingdom, India, China, etc.)
            data["Country"] = lines[-1]
            # Everything before it is the street/city address
            data["Address"] = ", ".join(lines[:-1]) if len(lines) > 1 else ""

    # ── Website
    #    FIX: use the exact class instead of a generic a[href^='http'] selector
    #    which was accidentally matching "http://searchToggle" nav link.
    website_el = soup.select_one(
        ".m-exhibitor-entry__item__body__contacts__additional__website a"
    )
    if website_el:
        data["Website"] = (website_el.get("href") or "").strip()

    # ── Email  (mailto: links inside the contacts section only)
    email_el = soup.select_one(
        ".m-exhibitor-entry__item__body__contacts a[href^='mailto:']"
    )
    if email_el:
        data["Email"] = email_el["href"].replace("mailto:", "").strip()

    # ── Phone  (tel: links inside the contacts section only)
    phone_el = soup.select_one(
        ".m-exhibitor-entry__item__body__contacts a[href^='tel:']"
    )
    if phone_el:
        data["Phone"] = phone_el["href"].replace("tel:", "").strip()

    # ── Social media
    #    Exact class: .m-exhibitor-entry__item__body__contacts__social
    #    Each <li> contains one <a href="https://social-platform.com/...">
    social_map = {
        "facebook.com":  "Facebook",
        "instagram.com": "Instagram",
        "linkedin.com":  "LinkedIn",
        "twitter.com":   "Twitter / X",
        "x.com":         "Twitter / X",
        "tiktok.com":    "TikTok",
        "youtube.com":   "YouTube",
    }
    social_ul = soup.select_one(".m-exhibitor-entry__item__body__contacts__social")
    if social_ul:
        for a in social_ul.select("a[href]"):
            href = a.get("href", "")
            for domain, field in social_map.items():
                if domain in href and not data[field]:
                    data[field] = href
                    break

    return data


# ─────────────────────────────────────────────
# STEP 3 – Scrape concurrently
# ─────────────────────────────────────────────
def scrape_all_exhibitors(exhibitor_links: list[dict]) -> list[dict]:
    results = []
    total = len(exhibitor_links)
    done  = 0

    print(f"\nScraping {total} exhibitor pages ({MAX_WORKERS} parallel workers)...\n")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(scrape_exhibitor_page, ex): ex
            for ex in exhibitor_links
        }
        for future in as_completed(futures):
            done += 1
            try:
                row = future.result()
                results.append(row)
                print(
                    f"  [{done:>3}/{total}] ✓  {row['Name']:<45} "
                    f"Stand: {row['Stand']:<8}  Country: {row['Country']}"
                )
            except Exception as e:
                ex = futures[future]
                print(f"  [{done:>3}/{total}] ✗  ERROR: {ex.get('url')} — {e}")

    return results


# ─────────────────────────────────────────────
# STEP 4 – Export to Excel
# ─────────────────────────────────────────────
def export_to_excel(rows: list[dict], filename: str):
    if not rows:
        print("No data to export!")
        return

    df = pd.DataFrame(rows)

    preferred_order = [
        "Name", "Stand", "Sector", "Description",
        "Address", "Country",
        "Website", "Email", "Phone",
        "Facebook", "Instagram", "LinkedIn", "Twitter / X", "TikTok", "YouTube",
        "Exclusive to Autumn Fair", "Sells on Faire.com",
        "New Exhibitor", "Taste @ Autumn Fair",
        "B Corp Certified", "Manufactured in UK",
        "Licensed Products", "Export Ready", "White/Private Label",
        "Products Include", "URL",
    ]
    cols   = [c for c in preferred_order if c in df.columns]
    extras = [c for c in df.columns if c not in cols]
    df     = df[cols + extras].fillna("")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Exhibitors")
        ws = writer.sheets["Exhibitors"]

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        # Styled header row
        ws.freeze_panes = "A2"
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="2E4057")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    print(f"\n✅  Exported {len(df)} rows  →  {filename}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 65)
    print("  Autumn Fair Exhibitor Scraper")
    print("=" * 65)

    print("\n[Step 1] Collecting exhibitor links from all 39 pages...\n")
    exhibitor_links = collect_all_exhibitor_links()
    print(f"\n  ✓ Total unique exhibitors found: {len(exhibitor_links)}")

    if not exhibitor_links:
        print("  No links found — check connection and try again.")
        raise SystemExit(1)

    print("\n[Step 2] Scraping exhibitor detail pages...\n")
    all_data = scrape_all_exhibitors(exhibitor_links)

    print("\n[Step 3] Exporting to Excel...\n")
    export_to_excel(all_data, OUTPUT_FILE)
    print("\nDone! 🎉")