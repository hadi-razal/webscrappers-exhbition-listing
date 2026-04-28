"""
IFA Berlin 2025 - Full Exhibitor Scraper
Scrapes all ~1800 exhibitors across 90 pages
Uses concurrent threading to open multiple detail pages simultaneously
Outputs results to Excel (.xlsx)

Requirements:
    pip install requests beautifulsoup4 openpyxl lxml
"""

import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import random
import logging
import re
from urllib.parse import urljoin

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BASE_URL        = "https://www.ifa-berlin.com"
LISTING_URL     = "https://www.ifa-berlin.com/exhibitors"
TOTAL_PAGES     = 90
MAX_LIST_WORKERS = 5   # parallel page-list fetches
MAX_DETAIL_WORKERS = 10  # parallel exhibitor-detail fetches  ← opens many tabs at once
OUTPUT_FILE     = "ifa_exhibitors_2025.xlsx"
REQUEST_DELAY   = (0.5, 1.5)   # random sleep range between requests (seconds)
MAX_RETRIES     = 3

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.ifa-berlin.com/exhibitors",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# STEP 1 — Collect all exhibitor URLs
# ─────────────────────────────────────────────

def fetch_page(session: requests.Session, url: str, retries: int = MAX_RETRIES) -> BeautifulSoup | None:
    """Fetch a URL and return a BeautifulSoup object, with retry logic."""
    for attempt in range(1, retries + 1):
        try:
            time.sleep(random.uniform(*REQUEST_DELAY))
            resp = session.get(url, headers=HEADERS, timeout=20)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")
        except Exception as exc:
            log.warning(f"Attempt {attempt}/{retries} failed for {url}: {exc}")
            if attempt == retries:
                log.error(f"Giving up on {url}")
                return None
            time.sleep(attempt * 2)


def get_exhibitor_links_from_page(session: requests.Session, page_num: int) -> list[str]:
    """Return all exhibitor profile URLs from a listing page."""
    url = LISTING_URL if page_num == 1 else f"{LISTING_URL}?page={page_num}"
    soup = fetch_page(session, url)
    if not soup:
        return []

    links = set()
    for a in soup.select('a[href*="/exhibitors/"]'):
        href = a.get("href", "")
        # Filter out the listing page itself and anchors
        if href.startswith("/exhibitors/") and href != "/exhibitors/" and "#" not in href:
            full_url = urljoin(BASE_URL, href)
            links.add(full_url)

    log.info(f"  Page {page_num}: found {len(links)} exhibitor links")
    return list(links)


def collect_all_exhibitor_urls() -> list[str]:
    """Collect all exhibitor URLs across all pages using thread pool."""
    log.info(f"── Collecting exhibitor URLs from {TOTAL_PAGES} pages ──")
    all_urls: set[str] = set()

    with requests.Session() as session:
        with ThreadPoolExecutor(max_workers=MAX_LIST_WORKERS) as executor:
            futures = {
                executor.submit(get_exhibitor_links_from_page, session, page): page
                for page in range(1, TOTAL_PAGES + 1)
            }
            for future in as_completed(futures):
                urls = future.result()
                all_urls.update(urls)

    log.info(f"Total exhibitor URLs collected: {len(all_urls)}")
    return sorted(all_urls)


# ─────────────────────────────────────────────
# STEP 2 — Scrape each exhibitor detail page
# ─────────────────────────────────────────────

SOCIAL_PATTERNS = {
    "instagram":  ["instagram.com"],
    "youtube":    ["youtube.com"],
    "facebook":   ["facebook.com"],
    "linkedin":   ["linkedin.com"],
    "twitter_x":  ["twitter.com", "x.com"],
    "tiktok":     ["tiktok.com"],
}


def classify_link(href: str) -> str:
    """Return the social platform name or 'website'."""
    for platform, patterns in SOCIAL_PATTERNS.items():
        if any(p in href for p in patterns):
            return platform
    return "website"


def scrape_exhibitor(session: requests.Session, url: str) -> dict:
    """Scrape a single exhibitor detail page and return a data dict."""
    data = {
        "name":             "",
        "slug":             url.rstrip("/").split("/")[-1],
        "exhibitor_url":    url,
        "category":         "",
        "hall":             "",
        "stand":            "",
        "country":          "",
        "website":          "",
        "email":            "",
        "instagram":        "",
        "youtube":          "",
        "facebook":         "",
        "linkedin":         "",
        "twitter_x":        "",
        "tiktok":           "",
        "logo_url":         "",
        "description":      "",
        "team_contact_email": "",
        "team_contact_linkedin": "",
        "ifa_moments":      "",
        "featured_products": "",
        "products_count":   0,
    }

    soup = fetch_page(session, url)
    if not soup:
        data["name"] = f"[ERROR fetching {url}]"
        return data

    # ── Basic info ──────────────────────────────────
    h1 = soup.select_one("h1")
    if h1:
        data["name"] = h1.get_text(strip=True)

    cat = soup.select_one(".chip.show-area")
    if cat:
        data["category"] = cat.get_text(strip=True)

    hall_el = soup.select_one(".brand-location-hall")
    if hall_el:
        data["hall"] = hall_el.get_text(strip=True)

    stand_el = soup.select_one(".brand-location-stand")
    if stand_el:
        data["stand"] = stand_el.get_text(strip=True)

    country_el = soup.select_one(".brand-detail small")
    if country_el:
        data["country"] = country_el.get_text(strip=True)

    # ── Logo ────────────────────────────────────────
    logo = soup.select_one(".brand-detail img, .brand-image img")
    if logo and logo.get("src"):
        data["logo_url"] = logo["src"]

    # ── Social / Website links ───────────────────────
    header_links = soup.select(".brand-detail-header-texts a[href], .brand-social a[href]")
    website_candidates = []
    for a in header_links:
        href = a.get("href", "").strip()
        if not href or "ifa-berlin.com" in href:
            continue
        if href.startswith("mailto:"):
            data["email"] = href.replace("mailto:", "").strip()
            continue
        platform = classify_link(href)
        if platform == "website":
            website_candidates.append(href)
        elif not data[platform]:           # keep first found
            data[platform] = href

    if website_candidates:
        data["website"] = website_candidates[0]

    # ── Description ─────────────────────────────────
    profile_section = soup.select_one("#profile")
    if profile_section:
        # Remove the header card portion
        for el in profile_section.select(".brand-detail"):
            el.decompose()
        raw_text = profile_section.get_text(separator="\n", strip=True)
        lines = [l.strip() for l in raw_text.splitlines() if l.strip() and len(l.strip()) > 15]
        data["description"] = "\n".join(lines)

    # ── Team contacts ───────────────────────────────
    contact_emails = []
    contact_linkedins = []

    # Method 1: dedicated .team-contact elements
    for contact in soup.select(".team-contact"):
        mail_a = contact.select_one('a[href^="mailto:"]')
        if mail_a:
            contact_emails.append(mail_a["href"].replace("mailto:", "").strip())
        li_a = contact.select_one('a[href*="linkedin"]')
        if li_a:
            contact_linkedins.append(li_a["href"].strip())

    # Method 2: any mailto on page (excluding IFA management)
    if not contact_emails:
        for a in soup.select('a[href^="mailto:"]'):
            email_addr = a["href"].replace("mailto:", "").strip()
            if "ifa-management" not in email_addr:
                contact_emails.append(email_addr)

    data["team_contact_email"]    = " | ".join(dict.fromkeys(contact_emails))
    data["team_contact_linkedin"] = " | ".join(dict.fromkeys(contact_linkedins))

    # ── IFA Moments / Events ────────────────────────
    moments = []
    for event in soup.select(".programme-event"):
        date_el  = event.select_one(".programme-event-date")
        title_el = event.select_one(".programme-event-title")
        loc_el   = event.select_one(".programme-event-location")
        dur_el   = event.select_one(".programme-event-duration")

        parts = []
        if date_el:  parts.append(date_el.get_text(strip=True))
        if title_el: parts.append(title_el.get_text(strip=True))
        if loc_el:   parts.append(loc_el.get_text(strip=True))
        if dur_el:   parts.append(dur_el.get_text(strip=True))
        if parts:
            moments.append(" | ".join(parts))

    data["ifa_moments"] = "\n".join(moments)

    # ── Featured Products ───────────────────────────
    products = []
    for prod in soup.select(".product"):
        name_el = prod.select_one(".product-name")
        link_el = prod.select_one(".product-link")
        abs_el  = prod.select_one(".product-abstract")

        pname = name_el.get_text(strip=True) if name_el else ""
        plink = ""
        if link_el:
            a_tag = link_el.select_one("a")
            plink = a_tag["href"] if a_tag else link_el.get_text(strip=True)
        pdesc = abs_el.get_text(strip=True)[:200] if abs_el else ""

        if pname:
            products.append(f"{pname} | {plink} | {pdesc}" if plink else pname)

    data["products_count"]    = len(products)
    data["featured_products"] = "\n\n".join(products)

    log.info(f"  ✓ {data['name']} ({data['country']}) — "
             f"{data['products_count']} products, "
             f"{len(moments)} events")
    return data


def scrape_all_exhibitors(urls: list[str]) -> list[dict]:
    """Scrape all exhibitor pages concurrently (simulates many open tabs)."""
    log.info(f"── Scraping {len(urls)} exhibitor pages with {MAX_DETAIL_WORKERS} parallel workers ──")
    results = []

    with requests.Session() as session:
        session.headers.update(HEADERS)
        with ThreadPoolExecutor(max_workers=MAX_DETAIL_WORKERS) as executor:
            futures = {executor.submit(scrape_exhibitor, session, url): url for url in urls}
            for future in as_completed(futures):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as exc:
                    log.error(f"Unexpected error: {exc}")

    return results


# ─────────────────────────────────────────────
# STEP 3 — Write to Excel
# ─────────────────────────────────────────────

COLUMNS = [
    ("Company Name",            "name"),
    ("Exhibitor Page URL",      "exhibitor_url"),
    ("Category / Show Area",    "category"),
    ("Hall",                    "hall"),
    ("Stand",                   "stand"),
    ("Country",                 "country"),
    ("Website",                 "website"),
    ("Email",                   "email"),
    ("Instagram",               "instagram"),
    ("YouTube",                 "youtube"),
    ("Facebook",                "facebook"),
    ("LinkedIn",                "linkedin"),
    ("Twitter / X",             "twitter_x"),
    ("TikTok",                  "tiktok"),
    ("Team Contact Email",      "team_contact_email"),
    ("Team Contact LinkedIn",   "team_contact_linkedin"),
    ("Logo URL",                "logo_url"),
    ("Description",             "description"),
    ("IFA Moments (Events)",    "ifa_moments"),
    ("Products Count",          "products_count"),
    ("Featured Products",       "featured_products"),
    ("Slug",                    "slug"),
]

HEADER_COLOR = "1A1A2E"   # dark navy
ALT_ROW_COLOR = "F0F4FF"  # light blue tint


def write_excel(data: list[dict], filename: str) -> None:
    log.info(f"── Writing {len(data)} records to {filename} ──")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IFA Exhibitors 2025"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    alt_fill     = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    wrap_align   = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")

    # Write header row
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    # Column widths (approximate)
    col_widths = {
        "Company Name":           30,
        "Exhibitor Page URL":     45,
        "Category / Show Area":   22,
        "Hall":                   10,
        "Stand":                  12,
        "Country":                18,
        "Website":                35,
        "Email":                  35,
        "Instagram":              35,
        "YouTube":                35,
        "Facebook":               35,
        "LinkedIn":               35,
        "Twitter / X":            30,
        "TikTok":                 30,
        "Team Contact Email":     35,
        "Team Contact LinkedIn":  35,
        "Logo URL":               40,
        "Description":            60,
        "IFA Moments (Events)":   50,
        "Products Count":         14,
        "Featured Products":      60,
        "Slug":                   28,
    }
    for idx, (col_name, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(idx)
        ].width = col_widths.get(col_name, 20)

    # Sort: alphabetically by company name
    data_sorted = sorted(data, key=lambda d: d.get("name", "").lower())

    # Write data rows
    for row_idx, record in enumerate(data_sorted, start=2):
        row_data = [record.get(field, "") for _, field in COLUMNS]
        ws.append(row_data)
        row = ws[row_idx]

        # Alternate row fill
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_fill

        # Wrap text for all cells + top-align
        for cell in row:
            cell.alignment = wrap_align

        # Set max row height for description / products rows
        ws.row_dimensions[row_idx].height = 60

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    log.info(f"✅ Saved: {filename}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    start = time.time()
    log.info("═══════════════════════════════════════════")
    log.info("  IFA Berlin 2025 — Exhibitor Scraper")
    log.info("═══════════════════════════════════════════")

    # Step 1 — get all exhibitor URLs
    exhibitor_urls = collect_all_exhibitor_urls()

    if not exhibitor_urls:
        log.error("No exhibitor URLs found. Exiting.")
        return

    # Step 2 — scrape all detail pages concurrently
    all_data = scrape_all_exhibitors(exhibitor_urls)

    # Step 3 — write to Excel
    write_excel(all_data, OUTPUT_FILE)

    elapsed = time.time() - start
    log.info(f"Done! {len(all_data)} exhibitors scraped in {elapsed:.0f}s")
    log.info(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()